# -*- coding: utf-8 -*-
import os
import sys
import time
import json
import re
import hashlib
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
load_dotenv()

CSV_INPUT  = "sample_products.csv"
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "captions_output.xlsx"
GROQ_URL   = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "llama-3.3-70b-versatile"
MAX_RETRIES  = 3
RETRY_DELAY  = 15

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
MODE = "groq" if GROQ_API_KEY else "template"

# ── Excel styles ──────────────────────────────────────────────────────────────
PINK_FILL       = PatternFill(start_color="F472B6", end_color="F472B6", fill_type="solid")
LIGHT_PINK_FILL = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")
WHITE_FILL      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GRAY_FILL       = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
HEADER_FONT     = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
LABEL_FONT      = Font(name="Calibri", bold=True, size=11, color="BE185D")
VALUE_FONT      = Font(name="Calibri", size=11, color="1F2937")
THIN_BORDER     = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# ── Lookup maps ───────────────────────────────────────────────────────────────
POST_TIME_MAP = {
    "instagram": "7h-9h sang hoac 19h-21h toi (moi truong VN)",
    "facebook":  "11h-13h trua hoac 20h-22h toi",
    "tiktok":    "17h-19h chieu hoac 21h-23h toi",
}

TONE_DESC_MAP = {
    "luxury":    "sang trong, tinh te, cao cap",
    "friendly":  "gan gui, than thien, de tiep can",
    "energetic": "nang dong, manh me, kich thich hanh dong",
}

PLATFORM_DESC_MAP = {
    "instagram": "Instagram (visual-first, nhieu hashtag, caption suc tich)",
    "facebook":  "Facebook (storytelling dai hon, CTA ro rang)",
    "tiktok":    "TikTok (hook dau manh, ngan, emoji nhieu, trend-aware)",
}

EMOJI_MAP = {
    ("skincare", "friendly"):  ["✨", "💕", "🌿", "💧", "🧴"],
    ("skincare", "luxury"):    ["💎", "✨", "🌸", "👑", "💫"],
    ("skincare", "energetic"): ["🔥", "⚡", "💪", "🎯", "✅"],
    ("makeup",   "friendly"):  ["💄", "✨", "💕", "🌟", "🎀"],
    ("makeup",   "luxury"):    ["💄", "💎", "✨", "👑", "🌹"],
    ("makeup",   "energetic"): ["💄", "🔥", "✨", "💥", "🎯"],
}

INGREDIENT_BENEFIT_MAP = {
    "snail":        "phuc hoi va tai tao da chuyen sau",
    "hyaluronic":   "cap am sau va cang bong da tuc thi",
    "niacinamide":  "mo tham va thu nho lo chan long",
    "aha":          "tay te bao chet nhe nhang va lam sang da",
    "bha":          "thong thoang lo chan long va kiem soat mun",
    "vitamin c":    "lam sang va chong oxy hoa hieu qua",
    "green tea":    "chong oxy hoa va kiem soat ba nhon tu nhien",
    "ceramide":     "phuc hoi hang rao bao ve da ben vung",
    "panthenol":    "lam diu va phuc hoi da ton thuong",
    "peptide":      "kich thich collagen va chong lao hoa",
    "berry":        "duong am va lam mem moi suot dem",
    "madecassoside":"lam diu kich ung va tai tao da",
    "heartleaf":    "khang viem va lam diu da nhay cam",
    "tea tree":     "khang khuan tu nhien va giam mun",
    "amino acid":   "nuoi duong va tang do dan hoi da",
    "seed oil":     "duong am sau va lam mem min da",
    "allantoin":    "lam diu da va ho tro tai tao te bao",
}


# ── Helper functions ──────────────────────────────────────────────────────────

def product_seed(name: str, n: int) -> int:
    return int(hashlib.md5(name.encode()).hexdigest(), 16) % n


def parse_ingredients(raw: str) -> tuple[str, str]:
    parts = [p.strip() for p in raw.split("-")]
    i1 = parts[0] if parts else raw
    i2 = parts[1] if len(parts) > 1 else i1
    return i1, i2


def get_benefit(raw: str) -> str:
    low = raw.lower()
    for key, val in INGREDIENT_BENEFIT_MAP.items():
        if key in low:
            return val
    return "cai thien lan da ro ret"


def get_audience_short(audience: str) -> str:
    cleaned = re.sub(r"^[Nn][uư][̃]?\s*\d+[-–]\d+\s*tu[oô]i\s*", "", audience).strip()
    return cleaned if cleaned else audience


def get_emojis(category: str, tone: str) -> list[str]:
    return EMOJI_MAP.get((category.lower(), tone.lower()), ["✨", "💕", "🌿", "💧", "🧴"])


def count_words(text: str) -> int:
    return len(text.split())


def build_hashtags(product: dict) -> str:
    name = product["product_name"]
    category = product.get("category", "skincare")
    platform = product.get("platform", "instagram")
    ingredients = product.get("key_ingredients", "")

    n = name.lower()
    if "cosrx" in n:       brand = ["#COSRX", "#COSRXVietnam"]
    elif "laneige" in n:   brand = ["#Laneige", "#LaneigeVN"]
    elif "some by mi" in n:brand = ["#SomeByMi", "#SomeByMiVN"]
    elif "innisfree" in n: brand = ["#Innisfree", "#InnisfreeVN"]
    elif "etude" in n:     brand = ["#EtudeHouse", "#EtudeVN"]
    else:                  brand = ["#kbeauty", "#kbeautyVN"]

    cat = {
        "skincare": ["#skincare", "#chamsocda", "#skincareroutine"],
        "makeup":   ["#makeup", "#trangdiem", "#makeuptips"],
    }.get(category.lower(), ["#beauty", "#lamdep"])

    plat = {
        "instagram": ["#skinstaagram", "#beautygram", "#glowingskin"],
        "tiktok":    ["#skintok", "#kbeautytiktok", "#beautytips"],
        "facebook":  ["#reviewmypham", "#myphamhanquoc", "#skincarevn"],
    }.get(platform.lower(), ["#beauty"])

    il = ingredients.lower()
    ing = []
    if "snail" in il:                              ing.append("#snailmucin")
    if "green tea" in il:                          ing.append("#greenteaskincare")
    if "niacinamide" in il:                        ing.append("#niacinamide")
    if "aha" in il or "bha" in il:                 ing.append("#exfoliant")
    if "hyaluronic" in il:                         ing.append("#hyaluronicacid")
    if "panthenol" in il:                          ing.append("#panthenol")
    if "vitamin c" in il:                          ing.append("#vitamincskincare")
    if "berry" in il:                              ing.append("#lipcare")
    if "peptide" in il:                            ing.append("#peptide")
    if "madecassoside" in il or "heartleaf" in il: ing.append("#calmskin")
    if not ing:                                    ing = ["#skincarescience"]

    general = ["#myphamhanquoc", "#lamdep", "#beautyvietnam", "#skingoals", "#kbeauty"]
    all_tags = brand[:2] + cat[:2] + plat[:2] + ing[:2] + general[:3]
    return " ".join(list(dict.fromkeys(all_tags))[:12])


# ── Template data ─────────────────────────────────────────────────────────────

TEMPLATES = {
    "friendly": {
        "c1": [
            "Minh vua thu {name} va da da thay doi hoan toan! Nho {i1} va {i2}, lan da minh am muot hon, khong con kho cang sau rua mat. Ban dang tim skincare nhe nhang? Day roi nhe!",
            "Gioi thieu em cung moi trong routine — {name}! {i1} giup {benefit}, hoan toan phu hop voi {audience}. Da dung 2 tuan va khong the dung lai duoc!",
            "Da dang gap van de? {name} voi {i1} la giai phap minh tim thay! Nhe nhang, hieu qua va khong gay kich ung. Skincare don gian ma hieu qua — chinh xac la day!",
            "Routine sang cua minh gio khong the thieu {name}! {i1} lam da minh {benefit}. Ai dang tim san pham diu nhe cho {audience} thi recommend ngay day nhe!",
        ],
        "c2": [
            "Minh da thu qua rat nhieu san pham skincare truoc khi tim ra {name} — va day thuc su la nguoi ban dong hanh ma lan da minh can.\n\n{i1} duoc chiet xuat tu nhien giup {benefit}, trong khi {i2} bo sung do am va lam diu da tuc thi. Ket qua sau 2 tuan: da mem hon, it do hon va trong khoe khoang hon nhieu.\n\nBan dang dung gi cho {audience}? Thu {name} di va ke minh nghe ket qua nhe!",
            "Skincare khong phai cu dat tien moi tot — minh hoc duoc dieu nay sau khi dung {name}!\n\nSan pham nay chua {i1} duoc dermatologist khuyen dung cho {audience}, ket hop voi {i2} de toi uu hieu qua duong da. Chi can vai buoc don gian moi ngay la da da cam nhan duoc su khac biet.\n\nBan da thu chua? Tag ban be de cung cham da nao!",
            "Hom nay minh muon review that long ve {name} sau 1 thang su dung!\n\nDieu minh thich nhat: {i1} tham nhanh, khong nhon rit, va thuc su giup da {benefit}. {i2} trong cong thuc con giup lam diu nhung ngay da 'kho o' nhat.\n\nPhu hop voi {audience} — minh da thay ket qua ro rang va se tiep tuc dung. Ai co cau hoi cu inbox minh nhe!",
            "Ban oi, minh can phai chia se ngay ve {name} vi da minh dang o trang thai tot nhat tu truoc den nay!\n\n{i1} la thanh phan chinh duoc chung minh lam sang to giup {benefit} mot cach tu nhien va ben vung. Ket hop voi {i2}, san pham tao thanh lop bao ve hoan hao cho da moi ngay.\n\nAi dang trong giai doan tim skincare phu hop — thu {name} di! Minh san sang tu van qua inbox.",
        ],
        "c3": [
            "Chao cac ban! Hom nay minh se review chi tiet ve {name} — san pham minh da dung suot 6 tuan qua!\n\nLy do minh chon san pham nay: da minh thuoc loai {audience} nen can san pham vua hieu qua vua khong gay kich ung. Va {name} da vuot qua moi ky vong!\n\nThanh phan noi bat:\n- {i1}: giup {benefit}\n- {i2}: tang cuong do am va lam diu da\n\nKet qua sau 6 tuan:\n- Tuan 1-2: da bat dau am hon, it cang kho\n- Tuan 3-4: tone da deu hon, bot do ro rang\n- Tuan 5-6: da cang bong, khoe khoang tu ben trong\n\nCach dung cua minh: thoa 3-4 giot sau buoc toner, massage nhe nhang va de tham tu nhien. Dung sang va toi deu on!\n\nAi can mua cung thi comment nhe, minh gom don! Link order trong bio.\n\n{hashtags}",
            "{name} — tai sao minh khong dung som hon nhi!\n\nMinh kha ken chon skincare vi {audience}. Nhung {name} da thay doi hoan toan dinh nghia ve san pham duong da 'phu hop'.\n\nDiem manh cua san pham:\n- {i1}: thanh phan nay giup {benefit} mot cach tu nhien va ben vung\n- {i2}: bo tro hoan hao, tang do tham va hieu qua tong the\n- Khong chua con, paraben — an toan cho da nhay cam\n\nReview that su:\n- Tuan dau: da hoi la voi san pham moi nhung khong co phan ung tieu cuc nao\n- Tuan 2 tro di: da am muot, tone deu hon va khong noi mun them!\n\nAi dang trong giai doan tim skincare phu hop — thu {name} di! Minh san sang tu van qua inbox.\n\n{hashtags}",
        ],
    },
    "luxury": {
        "c1": [
            "Khi lan da xung dang duoc cham soc boi nhung gi tinh tuy nhat. {name} voi {i1} thuan khiet — moi giot la mot loi hua ve lan da {benefit}. Sang trong khong can giai thich.",
            "Tinh hoa skincare Han Quoc trong tung giot tinh chat. {name} — {i1} duoc chiet xuat theo cong nghe doc quyen, mang den lan da {benefit} ma ban mo uoc. Dang cap dich thuc.",
            "Bo suu tap skincare chua hoan chinh neu thieu {name}. {i1} tinh che cao cap ket hop {i2} tao nen cong thuc {benefit} hoan hao. Mot lan thu — mai mai trung thanh.",
            "Khong phai ngau nhien ma {name} duoc tin dung boi nhung ai hieu gia tri cua lan da khoe. {i1} — thanh phan vang giup {benefit} tung ngay. Trai nghiem dang cap ngay hom nay.",
        ],
        "c2": [
            "Lan da hoan hao khong phai la may man — do la ket qua cua viec lua chon dung san pham.\n\n{name} duoc tao ra cho nhung nguoi tran trong chat luong. {i1} tinh khiet nhat duoc chiet xuat va ket hop cung {i2} theo ti le vang, tao ra hieu qua {benefit} vuot troi.\n\nDanh cho {audience} — nhung nguoi hieu rang lan da la tam guong phan chieu suc khoe va phong cach song. Link dat hang trong bio.",
            "Trong the gioi skincare day ap lua chon, {name} dung o mot dang cap rieng biet.\n\nCong thuc doc quyen voi {i1} — khong don gian chi la duong da ma la mot nghi thuc cham soc ban than cao cap. {i2} ket hop hai hoa de {benefit}, mang lai ket qua ban co the cam nhan ngay tu lan su dung dau tien.\n\nBoi vi ban xung dang duoc huong nhung dieu tot nhat. DM de duoc tu van.",
            "Skincare luxury khong phai la xa xi — do la su dau tu thong minh cho tuong lai cua lan da ban.\n\n{name} voi {i1} tinh khiet la minh chung cho dieu do. Moi phan tu {i1} duoc kiem nghiem nghiem ngat truoc khi den tay ban, dam bao {benefit} toi uu nhat.\n\nDanh rieng cho {audience} — nhung nguoi khong thoa hiep voi chat luong. So luong co han moi dot nhap.",
            "Huong thom nhe nhang, ket cau min mang, hieu qua sau sac — do la {name}.\n\n{i1} duoc lay tu nguon nguyen lieu tu nhien chon loc, ket hop voi {i2} trong phong lab hien dai nhat Han Quoc. Ket qua? Lan da {benefit} theo cach tu nhien va ben vung nhat.\n\nPhu hop voi {audience} tim kiem giai phap skincare dang cap. Inbox de duoc goi qua dac biet.",
        ],
        "c3": [
            "Hanh trinh tim kiem lan da hoan hao ket thuc tai day — voi {name}.\n\nKhong phai ngau nhien ma cac beauty editor Han Quoc xep {name} vao danh sach must-have nhieu nam lien tiep. Day la ly do:\n\nThanh phan dang cap:\n- {i1}: Duoc chiet xuat theo quy trinh kiem soat nhiet do nghiem ngat, bao toan 99% hoat chat giup {benefit}\n- {i2}: Thanh phan bo tro duoc phat trien boi doi ngu dermatologist hang dau Han Quoc\n\nKhoa hoc dang sau su hoan hao:\nCong thuc cua {name} duoc kiem nghiem lam sang tren 500 tinh nguyen vien trong 12 tuan. Ket qua: 94% nhan thay da {benefit} sau 4 tuan su dung deu dan.\n\nCach dung dung chuan:\nSau toner, thoa 2-3 giot len da, vo nhe de tham. Dung sang va toi. Ket hop kem chong nang vao buoi sang de toi uu hieu qua.\n\nDanh cho {audience} — nhung nguoi hieu rang dau tu vao lan da la dau tu sinh loi nhat. Link order trong bio. Free shipping cho don tu 500k.\n\n{hashtags}",
            "Khi toi gian gap hieu qua — {name} la cau tra loi.\n\nMoi san pham trong dong nay deu trai qua 3 nam nghien cuu va phat trien tai Han Quoc truoc khi ra mat. Voi {name}, dieu do co nghia la:\n\nCong nghe chiet xuat {i1}:\nKhong phai {i1} nao cung nhu nhau. {name} dung {i1} duoc xu ly o nhiet do thap de bao toan toan bo hoat chat — toi uu hon 3 lan so voi phuong phap thong thuong.\n\nCong thuc synergy voi {i2}:\n{i2} khong chi la thanh phan phu — no khuech dai tac dung cua {i1} va dam bao {benefit} keo dai suot ngay.\n\nKet qua thuc te:\n89% nguoi dung thay da am hon sau 7 ngay\n76% ghi nhan da deu mau sau 4 tuan\n0% bao cao kich ung hoac tac lo chan long\n\nDanh cho {audience} muon ket qua that su. Inbox de duoc tu van skincare routine phu hop.\n\n{hashtags}",
        ],
    },
    "energetic": {
        "c1": [
            "DUNG LAI! {name} vua thay doi toan bo skincare game cua minh! {i1} cong {i2} — combo cuc manh giup {benefit} chi trong 30 ngay. Ai dang chiu dung da xau? Day la giai phap!",
            "KHONG THE TIN DUOC! {name} voi {i1} da lam duoc dieu ma ca chuc san pham truoc khong lam duoc! Da minh gio: {benefit}. Ai can chi tiet? COMMENT ngay!",
            "Da dang hanh ban? {name} la cau tra loi! {i1} hoat dong 24/7 de {benefit}. Duoc recommend boi cong dong skincare toan cau. ORDER truoc khi het hang!",
            "VIRAL vi ly do xung dang! {name} voi {i1} dang gay bao skincare community! Minh da thu va KET QUA THAT khong can filter. {benefit} — khong phai quang cao!",
        ],
        "c2": [
            "Nam ngoai da minh la tham hoa. Nam nay? Hoan toan khac nho {name}!\n\n{i1} trong {name} la bi mat ma cac dermatologist biet nhung it nguoi chia se. Hoat chat nay {benefit} — duoc chung minh boi hang nghin case study tren toan the gioi. {i2} con tang cuong them hieu qua, giam thoi gian cho ket qua xuong con mot nua.\n\nDanh cho {audience} — ban khong can phai chiu dung them nua! TAG ban be can biet va ORDER ngay truoc khi het hang nhe!",
            "Minh se noi thang: {name} la san pham tot nhat minh tung dung cho {audience}!\n\nTai sao? Vi {i1} khong chi 'duong' ma con chua lanh da tu ben trong. {i2} ho tro them de {benefit} nhanh hon, ro hon. Ket qua minh thay sau 30 ngay khien minh khong the khong share!\n\nSale dang chay — giam 25% hom nay. Link trong bio. SAVE post nay lai de con mua!",
            "3 ly do ban CAN {name} ngay bay gio:\n\n1. {i1} — hoat chat so 1 duoc dermatologist khuyen dung cho {audience}\n2. {i2} — tang cuong hieu qua gap doi so voi dung rieng le\n3. Ket qua {benefit} duoc chung minh sau 4 tuan\n\nMinh da thu va ket qua vuot xa ky vong! Comment MUA de minh gui link cong voucher 15% cho ban nhe!",
            "OK minh phai noi ra vi qua excited: {name} vua cuu da minh!\n\nCon nho da minh hoi thang truoc? {audience} ma khong tim dung san pham thi kho lam. Nhung {i1} trong {name} da lam dieu ky dieu — {benefit} chi sau 3 tuan! {i2} con lam minh ngac nhien them voi viec da khoe tu ben trong.\n\nShare ngay cho ban be dang can! Link order trong bio — FREESHIP hom nay!",
        ],
        "c3": [
            "30 NGAY THACH THUC voi {name} — va day la ket qua CUC SOC!\n\nMinh biet nhieu nguoi hoai nghi ve skincare viral, nhung {name} thuc su la ngoai le. Hay de minh breakdown tai sao:\n\nCong thuc KHONG THE BO QUA:\n- {i1}: Hoat chat chu luc — duoc chung minh {benefit} chi trong 2-4 tuan. Khong phai placebo, khong phai marketing — co nghien cuu khoa hoc backup!\n- {i2}: Sidekick hoan hao — tang do tham thau va keo dai hieu qua cua {i1} gap 2 lan\n\nTIMELINE cua minh:\n- Ngay 1-7: da hoi quen voi thanh phan moi, khong co purging\n- Ngay 8-14: da bat dau am hon, bot bong dau\n- Ngay 15-21: tone da deu hon, lo chan long trong nho hon\n- Ngay 22-30: {benefit} — KHONG FILTER, KHONG EDIT!\n\nHANH DONG NGAY: Chi con hang co han trong kho. Link order trong bio. Nhap code SKIN15 de giam them 15%. Gom don cung minh!\n\n{hashtags}",
            "Cong dong skincare dang NOI GI ve {name}? Minh tong hop ngay day!\n\n'Dung 2 tuan da minh doi han' — nguoi dung A\n'Khong ngo {i1} lai hieu qua den vay' — nguoi dung B\n'Cuoi cung tim duoc san pham phu hop voi {audience}' — nguoi dung C\n\nVA DAY LA LY DO KHOA HOC:\n{i1} tac dong truc tiep vao co che {benefit} cua da. Khong phai ngau nhien ma hang trieu nguoi Chau A chon {name} lam staple trong routine.\n\n{i2} bo sung them lop bao ve va nuoi duong da tu ben trong — combo nay chinh la ly do san pham luon sold out!\n\nCANH BAO: San pham nay co the khien ban nghien skincare (minh da bi roi)!\n\nORDER: Link trong bio. Comment INFO de minh gui bang gia va cach dung dung nhat. SAVE de khong quen!\n\n{hashtags}",
        ],
    },
}


# ── Template caption generator ────────────────────────────────────────────────

def generate_template_captions(product: dict) -> dict:
    name     = product["product_name"]
    tone     = product.get("tone", "friendly").lower()
    platform = product.get("platform", "instagram").lower()
    category = product.get("category", "skincare").lower()
    i1, i2   = parse_ingredients(product.get("key_ingredients", ""))
    benefit  = get_benefit(product.get("key_ingredients", ""))
    audience = get_audience_short(product.get("target_audience", ""))
    emojis   = get_emojis(category, tone)
    hashtags = build_hashtags(product)
    post_time = POST_TIME_MAP.get(platform, "7h-9h sang hoac 19h-21h toi")

    pool = TEMPLATES.get(tone, TEMPLATES["friendly"])

    def pick(lst: list[str]) -> str:
        tpl = lst[product_seed(name, len(lst))]
        return tpl.format(
            name=name, i1=i1, i2=i2,
            benefit=benefit, audience=audience,
            hashtags=hashtags,
        )

    c1 = pick(pool["c1"])
    c2 = pick(pool["c2"])
    c3 = pick(pool["c3"])

    return {
        "caption_1": {"content": c1, "word_count": count_words(c1)},
        "caption_2": {"content": c2, "word_count": count_words(c2)},
        "caption_3": {"content": c3, "word_count": count_words(c3)},
        "emojis":    emojis,
        "post_time": post_time,
        "source":    "template",
    }


# ── Groq API ──────────────────────────────────────────────────────────────────

def build_groq_prompt(product: dict) -> str:
    return (
        "Ban la chuyen gia copywriting marketing my pham Han Quoc tai Viet Nam.\n"
        "Hay viet 3 bien the caption tieng Viet cho san pham sau:\n\n"
        f"Ten: {product['product_name']}\n"
        f"Danh muc: {product['category']}\n"
        f"Thanh phan chinh: {product['key_ingredients']}\n"
        f"Doi tuong: {product['target_audience']}\n"
        f"Giong van: {TONE_DESC_MAP.get(product.get('tone','friendly'), product.get('tone',''))}\n"
        f"Nen tang: {PLATFORM_DESC_MAP.get(product.get('platform','instagram'), product.get('platform',''))}\n\n"
        "Tra ve JSON hop le (khong them text nao khac ngoai JSON):\n\n"
        "{\n"
        '  "caption_1": {"content": "Caption ngan duoi 50 tu, hook manh", "word_count": 0},\n'
        '  "caption_2": {"content": "Caption trung binh 50-100 tu, storytelling", "word_count": 0},\n'
        '  "caption_3": {"content": "Caption dai 100+ tu, CTA ro rang, 10-12 hashtag", "word_count": 0},\n'
        '  "emojis": ["e1","e2","e3","e4","e5"],\n'
        '  "post_time": "Goi y thoi diem dang toi uu"\n'
        "}\n\n"
        "Luu y: viet tieng Viet tu nhien, nhan manh thanh phan thuc te, dien word_count chinh xac."
    )


def call_groq_api(product: dict) -> dict | None:
    try:
        import httpx
    except ImportError:
        print("    [!] httpx chua duoc cai. Dung template mode.")
        return None

    prompt  = build_groq_prompt(product)
    headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": GROQ_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 2000,
        "temperature": 0.75,
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with httpx.Client(timeout=60) as client:
                resp = client.post(GROQ_URL, headers=headers, json=payload)

            if resp.status_code == 429:
                if attempt < MAX_RETRIES:
                    print(f"    [!] Rate limit — cho {RETRY_DELAY}s ({attempt}/{MAX_RETRIES})")
                    time.sleep(RETRY_DELAY)
                    continue
                return None

            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"].strip()
            m = re.search(r"\{[\s\S]*\}", raw)
            if not m:
                continue
            result = json.loads(m.group())
            for k in ("caption_1", "caption_2", "caption_3"):
                if k in result and "content" in result[k]:
                    result[k]["word_count"] = count_words(result[k]["content"])
            result["source"] = "groq"
            return result

        except json.JSONDecodeError:
            if attempt == MAX_RETRIES:
                return None
        except Exception as e:
            print(f"    [x] Loi Groq API: {e}")
            return None

    return None


# ── Dispatcher ────────────────────────────────────────────────────────────────

def generate_captions(product: dict) -> dict:
    if MODE == "groq":
        result = call_groq_api(product)
        if result:
            return result
        print("    -> Fallback sang template mode cho san pham nay.")
    return generate_template_captions(product)


# ── Excel writers ─────────────────────────────────────────────────────────────

def write_product_sheet(wb: Workbook, product: dict, captions: dict, idx: int):
    name      = product["product_name"]
    safe_name = re.sub(r'[\[\]:*?/\\]', '', name)[:27]
    ws        = wb.create_sheet(title=f"{idx}. {safe_name}")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    row = 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value     = f"CAPTION MARKETING — {name.upper()}"
    c.fill      = PINK_FILL
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 36
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value     = "THONG TIN SAN PHAM"
    c.fill      = LIGHT_PINK_FILL
    c.font      = LABEL_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    source_label = "Nguon: Groq API (llama-3.3-70b)" if captions.get("source") == "groq" else "Nguon: Template Engine (offline)"
    info_rows = [
        ("Ten san pham",   product.get("product_name", "")),
        ("Danh muc",       product.get("category", "")),
        ("Thanh phan",     product.get("key_ingredients", "")),
        ("Doi tuong",      product.get("target_audience", "")),
        ("Giong van",      product.get("tone", "").capitalize()),
        ("Nen tang",       product.get("platform", "").capitalize()),
        ("Phuong thuc",    source_label),
    ]
    for label, value in info_rows:
        lc, vc = ws[f"A{row}"], ws[f"B{row}"]
        lc.value = label
        vc.value = value
        bg = GRAY_FILL if row % 2 == 0 else WHITE_FILL
        lc.fill = vc.fill = bg
        lc.font = Font(name="Calibri", bold=True, size=10, color="6B7280")
        vc.font = VALUE_FONT
        lc.alignment = Alignment(vertical="center", indent=1)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        lc.border = vc.border = THIN_BORDER
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    sections = [
        ("caption_1", "BIEN THE 1 — NGAN GON (Hook manh, duoi 50 tu)"),
        ("caption_2", "BIEN THE 2 — TRUNG BINH (Storytelling, 50-100 tu)"),
        ("caption_3", "BIEN THE 3 — DAY DU (Co hashtag & CTA, 100+ tu)"),
    ]
    for key, title in sections:
        data    = captions.get(key, {})
        content = data.get("content", "Khong co du lieu")
        wc      = data.get("word_count", 0)

        ws.merge_cells(f"A{row}:B{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.fill = LIGHT_PINK_FILL
        c.font = LABEL_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        lc, vc = ws[f"A{row}"], ws[f"B{row}"]
        lc.value = "Noi dung"
        lc.font  = Font(name="Calibri", bold=True, size=10, color="6B7280")
        lc.alignment = Alignment(vertical="top", indent=1)
        lc.border = THIN_BORDER
        vc.value = content
        vc.font  = VALUE_FONT
        vc.alignment = Alignment(vertical="top", wrap_text=True)
        vc.border = THIN_BORDER
        ws.row_dimensions[row].height = max(4, len(content) // 60 + 1) * 15
        row += 1

        lc, vc = ws[f"A{row}"], ws[f"B{row}"]
        lc.value = "So tu"
        lc.font  = Font(name="Calibri", bold=True, size=10, color="6B7280")
        lc.alignment = Alignment(vertical="center", indent=1)
        lc.border = THIN_BORDER
        vc.value = f"{wc} tu"
        vc.font  = VALUE_FONT
        vc.alignment = Alignment(vertical="center")
        vc.border = THIN_BORDER
        ws.row_dimensions[row].height = 20
        row += 2

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = "GOI Y BO SUNG"
    c.fill  = LIGHT_PINK_FILL
    c.font  = LABEL_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    row += 1

    emojis = captions.get("emojis", [])
    lc, vc = ws[f"A{row}"], ws[f"B{row}"]
    lc.value = "Emoji goi y"
    lc.font  = Font(name="Calibri", bold=True, size=10, color="6B7280")
    lc.alignment = Alignment(vertical="center", indent=1)
    lc.border = THIN_BORDER
    vc.value = "  ".join(emojis) if emojis else "-"
    vc.font  = Font(name="Segoe UI Emoji", size=14)
    vc.alignment = Alignment(vertical="center")
    vc.border = THIN_BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    lc, vc = ws[f"A{row}"], ws[f"B{row}"]
    lc.value = "Thoi diem dang"
    lc.font  = Font(name="Calibri", bold=True, size=10, color="6B7280")
    lc.alignment = Alignment(vertical="center", indent=1)
    lc.border = THIN_BORDER
    vc.value = captions.get("post_time", POST_TIME_MAP.get(product.get("platform", ""), "-"))
    vc.font  = VALUE_FONT
    vc.alignment = Alignment(vertical="center", wrap_text=True)
    vc.border = THIN_BORDER
    ws.row_dimensions[row].height = 20


def write_overview_sheet(wb: Workbook, products: list[dict], results: list[dict]):
    ws = wb.active
    ws.title = "Tong quan"

    headers = ["STT", "Ten san pham", "Danh muc", "Nen tang",
               "Giong van", "Nguon", "Trang thai", "So bien the"]
    ws.row_dimensions[1].height = 32
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PINK_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    for ci, w in enumerate([6, 40, 16, 14, 14, 20, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (prod, res) in enumerate(zip(products, results), 1):
        row  = ri + 1
        fill = WHITE_FILL if ri % 2 == 0 else GRAY_FILL
        src  = res.get("source", "template") if res else "-"
        ok   = res is not None
        data = [
            ri,
            prod.get("product_name", ""),
            prod.get("category", ""),
            prod.get("platform", ""),
            prod.get("tone", ""),
            "Groq API" if src == "groq" else "Template",
            "OK" if ok else "FAIL",
            3 if ok else 0,
        ]
        for ci, val in enumerate(data, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill
            c.font = VALUE_FONT
            c.alignment = Alignment(
                horizontal="left" if ci == 2 else "center",
                vertical="center", wrap_text=True,
            )
            c.border = THIN_BORDER
            if ci == 7:
                c.font = Font(name="Calibri", size=11,
                              color="16A34A" if ok else "DC2626", bold=True)
        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A2"


# ── Progress bar ──────────────────────────────────────────────────────────────

def print_progress(cur: int, total: int, name: str, ok: bool):
    filled = int(30 * cur / total)
    bar    = "#" * filled + "-" * (30 - filled)
    label  = "OK  " if ok else "SKIP"
    print(f"  [{bar}] {cur}/{total}  [{label}]  {name}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not Path(CSV_INPUT).exists():
        print(f"[ERROR] Khong tim thay file CSV: {CSV_INPUT}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 62)
    print("  CAPTION BATCH GENERATOR")
    if MODE == "groq":
        print("  Mode: GROQ API  (llama-3.3-70b-versatile) — mien phi")
    else:
        print("  Mode: TEMPLATE ENGINE  (offline, khong can API key)")
        print("  Tip: Dat GROQ_API_KEY=... trong .env de dung AI mode")
    print("=" * 62)

    df = pd.read_csv(CSV_INPUT)
    required = {"product_name", "category", "key_ingredients",
                "target_audience", "tone", "platform"}
    missing = required - set(df.columns)
    if missing:
        print(f"[ERROR] File CSV thieu cot: {missing}")
        sys.exit(1)

    products = df.to_dict(orient="records")
    total    = len(products)
    print(f"\n  Tim thay {total} san pham trong {CSV_INPUT}\n")

    results: list[dict] = []
    for i, prod in enumerate(products, 1):
        name = prod.get("product_name", f"San pham {i}")
        print(f"\n[{i}/{total}] {name}")
        print(f"  Tone: {prod.get('tone')} | Platform: {prod.get('platform')}")
        res = generate_captions(prod)
        results.append(res)
        print_progress(i, total, name, True)
        if MODE == "groq" and i < total:
            time.sleep(1)

    print("\n" + "=" * 62)
    print("  Dang xuat file Excel...")
    print("=" * 62)

    wb = Workbook()
    write_overview_sheet(wb, products, results)
    for si, (prod, res) in enumerate(zip(products, results), 1):
        write_product_sheet(wb, prod, res, si)

    wb.save(OUTPUT_FILE)
    print(f"\n  Xuat thanh cong: {total}/{total} san pham")
    print(f"  File da luu tai: {OUTPUT_FILE.resolve()}")
    print("\n" + "=" * 62 + "\n")


if __name__ == "__main__":
    main()
